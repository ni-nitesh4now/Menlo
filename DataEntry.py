from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl
import xlrd
from openpyxl import Workbook
import pathlib

root=Tk()
root.title("Data Entry")
root.geometry('700x400+300+200')
root.resizable(False,False)
root.configure(bg="#326273")

file=pathlib.Path('Data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="FullName"
    sheet['B1']="Phone Number"
    sheet['C1']="Age"
    sheet['D1']="Gender"
    sheet['E1']="Address"
    file.save('Data.xlsx')


#icon
icon_imae=PhotoImage(file="someth.png")
root.iconphoto(False,icon_imae)

#heading
Label(root,text="Please fill out this entry form:", font="arial 13",bg="#326273",fg="#fff").place(x=20,y=20)

#label
Label(root,text='Name',font=23,bg="#326273",fg="#fff").place(x=50,y=100)
Label(root,text='Contact No.',font=23,bg="#326273",fg="#fff").place(x=50,y=150)
Label(root,text='Age',font=23,bg="#326273",fg="#fff").place(x=50,y=200)
Label(root,text='Gender',font=23,bg="#326273",fg="#fff").place(x=370,y=200)
Label(root,text='Address',font=23,bg="#326273",fg="#fff").place(x=50,y=250)

#Entry
nameval=StringVar()
contactval=StringVar()
ageval=StringVar()

nameEntry=Entry(root,textvariable=nameval,width=45,bd=2,font=20)
nameEntry.place(x=200,y=100)

conEntry=Entry(root,textvariable=contactval,width=45,bd=2,font=20)
conEntry.place(x=200,y=150)

ageEntry=Entry(root,textvariable=ageval,width=15,bd=2,font=20)
ageEntry.place(x=200,y=200)

gender_comn=Combobox(root,values=['Male','Female','Other'],font='arial 14',state='r',width=14)
gender_comn.place(x=440,y=200)
gender_comn.set('Male')

addressEntry=Text(root,width=50,height=4,bd=2)
addressEntry.place(x=200,y=250)

#buttons
def Submit():
    name=nameval.get()
    contact=contactval.get()
    age=ageval.get()
    gender=gender_comn.get()
    addres=addressEntry.get(1.0,END)
    wb = openpyxl.load_workbook(filename='Data.xlsx')
    sheet=wb.active
    sheet.cell(column=1,row=sheet.max_row+1).value=name
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row).value=addres
    wb.save('Data.xlsx')
    messagebox.showinfo('info','Details added!')
    nameval.set(' ')
    ageval.set(' ')
    contactval.set(' ')
    addressEntry.delete(1.0,END)


def Clear():
    nameval.set('')
    ageval.set('')
    contactval.set('')
    addressEntry.delete(1.0,END)

Button(root,text="Submit",bg="#326273",fg='white', width=15,height=2,command=Submit).place(x=200,y=350)
Button(root,text="Clear",bg="#326273",fg='white', width=15,height=2,command=Clear).place(x=340,y=350)
Button(root,text="Exit",bg="#326273",fg='white', width=15,height=2,command=root.destroy).place(x=480,y=350)

root.mainloop()